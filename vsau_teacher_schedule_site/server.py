from __future__ import annotations

import hashlib
import io
import json
import mimetypes
import re
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    load_workbook = None
    get_column_letter = None

try:
    import xlrd
except ModuleNotFoundError:
    xlrd = None


ROOT_DIR = Path(__file__).resolve().parent
STATIC_DIR = ROOT_DIR / "static"
CACHE_DIR = ROOT_DIR / "data" / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
TREE_CACHE_FILE = ROOT_DIR / "data" / "tree-cache.json"
AGGREGATE_CACHE_DIR = ROOT_DIR / "data" / "aggregate-cache"
AGGREGATE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
LESSON_CACHE_DIR = ROOT_DIR / "data" / "lesson-cache"
LESSON_CACHE_DIR.mkdir(parents=True, exist_ok=True)

ROOT_FOLDER_ID = "1Gn8OEzbtxFBusuCnPCTBjQoK1AWKiVN2"
PARSER_VERSION = 10
TREE_CACHE_SECONDS = 10 * 60
FILE_CACHE_SECONDS = 5 * 60
AGGREGATE_CACHE_SECONDS = 6 * 60 * 60
POLL_SECONDS = 10 * 60
MAX_TREE_DEPTH = 6

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
)

tree_lock = threading.Lock()
tree_cache: dict[str, Any] = {"updated_at": 0.0, "data": None, "error": None}


class ApiError(Exception):
    def __init__(self, message: str, status: int = 500) -> None:
        super().__init__(message)
        self.status = status


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def fetch_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                charset = response.headers.get_content_charset() or "utf-8"
                return response.read().decode(charset, errors="replace")
        except Exception as exc:
            last_error = exc
            time.sleep(0.8)
    raise last_error or ApiError("Не удалось загрузить страницу Google Drive", 502)


def download_bytes(url: str) -> tuple[bytes, str]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    last_error: Exception | None = None
    for _ in range(3):
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                content_type = response.headers.get("content-type", "")
                return response.read(), content_type
        except Exception as exc:
            last_error = exc
            time.sleep(1)
    raise last_error or ApiError("Не удалось скачать файл Google Drive", 502)


def drive_embedded_url(folder_id: str) -> str:
    query = urllib.parse.urlencode({"id": folder_id})
    return f"https://drive.google.com/embeddedfolderview?{query}#list"


def drive_download_url(file_id: str) -> str:
    query = urllib.parse.urlencode({"export": "download", "id": file_id})
    return f"https://drive.google.com/uc?{query}"


def extract_drive_id(url: str) -> str:
    patterns = (
        r"/drive/folders/([^/?#]+)",
        r"/file/d/([^/?#]+)",
        r"[?&]id=([^&#]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return urllib.parse.unquote(match.group(1))
    raise ApiError(f"Не удалось найти id Google Drive в ссылке: {url}", 422)


def clean_html(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    return urllib.parse.unquote(value).replace("&amp;", "&").strip()


def parse_folder_entries(html: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    pattern = re.compile(
        r'<a href="(?P<href>[^"]+)"[^>]*>.*?'
        r'<div class="flip-entry-title">(?P<title>.*?)</div>.*?'
        r'<div class="flip-entry-last-modified"><div>(?P<modified>.*?)</div>',
        re.S,
    )

    for match in pattern.finditer(html):
        href = match.group("href").replace("&amp;", "&")
        title = clean_html(match.group("title"))
        modified = clean_html(match.group("modified"))
        if not title:
            continue

        is_folder = "/drive/folders/" in href
        item_id = extract_drive_id(href)
        extension = Path(title).suffix.lower().lstrip(".")
        entries.append(
            {
                "id": item_id,
                "title": title,
                "modified": modified,
                "type": "folder" if is_folder else "file",
                "extension": extension,
                "url": href,
                "downloadUrl": drive_download_url(item_id) if not is_folder else None,
                "viewUrl": href,
            }
        )

    return entries


def fetch_folder(folder_id: str) -> dict[str, Any]:
    html = fetch_text(drive_embedded_url(folder_id))
    title_match = re.search(r"<title>(.*?)</title>", html, re.S)
    title = clean_html(title_match.group(1)) if title_match else "Google Drive"
    return {
        "id": folder_id,
        "title": title.replace(" - Google Drive", "").replace(" – Dysk Google", ""),
        "type": "folder",
        "url": f"https://drive.google.com/drive/folders/{folder_id}",
        "children": parse_folder_entries(html),
    }


def build_tree(folder_id: str, depth: int = 0) -> dict[str, Any]:
    folder = fetch_folder(folder_id)
    if depth >= MAX_TREE_DEPTH:
        return folder

    children: list[dict[str, Any]] = list(folder["children"])
    folder_indexes = [index for index, child in enumerate(children) if child["type"] == "folder"]
    if folder_indexes:
        workers = min(8, len(folder_indexes))
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(build_tree, children[index]["id"], depth + 1): index
                for index in folder_indexes
            }
            for future in as_completed(futures):
                index = futures[future]
                child = children[index]
                try:
                    nested = future.result()
                    nested["title"] = child["title"]
                    nested["modified"] = child.get("modified")
                    children[index] = nested
                except Exception as exc:
                    child["error"] = str(exc)
                    children[index] = child

    folder["children"] = children
    return folder


def refresh_tree(force: bool = False) -> dict[str, Any]:
    with tree_lock:
        age = time.time() - tree_cache["updated_at"]
        if not force and tree_cache["data"] is not None and age < TREE_CACHE_SECONDS:
            return tree_cache["data"]
        if not force and tree_cache["data"] is None and TREE_CACHE_FILE.exists():
            try:
                data = json.loads(TREE_CACHE_FILE.read_text(encoding="utf-8"))
                tree_cache.update({"updated_at": time.time(), "data": data, "error": None})
                return data
            except Exception:
                pass

        try:
            data = build_tree(ROOT_FOLDER_ID)
            data["refreshedAt"] = now_iso()
            TREE_CACHE_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
            tree_cache.update({"updated_at": time.time(), "data": data, "error": None})
            return data
        except Exception as exc:
            tree_cache["error"] = str(exc)
            if tree_cache["data"] is not None:
                return tree_cache["data"]
            raise


def background_refresh() -> None:
    while True:
        try:
            refresh_tree(force=True)
        except Exception:
            pass
        time.sleep(POLL_SECONDS)


def cache_file_path(file_id: str, extension: str) -> Path:
    extension = extension or "bin"
    return CACHE_DIR / f"{file_id}.{extension}"


def get_cached_file(file_id: str, extension: str, force: bool = False) -> dict[str, Any]:
    path = cache_file_path(file_id, extension)
    age = time.time() - path.stat().st_mtime if path.exists() else float("inf")
    if force or not path.exists() or age > FILE_CACHE_SECONDS:
        data, content_type = download_bytes(drive_download_url(file_id))
        if b"<html" in data[:300].lower() and "text/html" in content_type:
            raise ApiError("Google Drive вернул HTML-страницу вместо файла. Проверьте публичный доступ.", 502)
        path.write_bytes(data)

    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "path": path,
        "size": path.stat().st_size,
        "sha256": digest,
        "cachedAt": datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat(timespec="seconds"),
    }


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def cell_style(cell: Any) -> dict[str, Any]:
    fill = getattr(cell.fill.fgColor, "rgb", None)
    fill = fill if isinstance(fill, str) else ""
    if fill and len(fill) == 8 and fill != "00000000":
        fill = "#" + fill[2:]
    else:
        fill = ""

    color = getattr(cell.font.color, "rgb", None)
    color = color if isinstance(color, str) else ""
    if color and len(color) == 8 and color != "00000000":
        color = "#" + color[2:]
    else:
        color = ""

    return {
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "fill": fill,
        "color": color,
        "align": cell.alignment.horizontal or "",
        "valign": cell.alignment.vertical or "",
    }


def parse_workbook(path: Path) -> dict[str, Any]:
    if load_workbook is None:
        raise ApiError("Не установлен пакет openpyxl. Установите: python -m pip install openpyxl", 500)

    source: Any = io.BytesIO(path.read_bytes()) if path.suffix.lower() == ".xls" else path
    workbook = load_workbook(source, data_only=True)
    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        populated: list[tuple[int, int]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if value_to_text(cell.value):
                    populated.append((cell.row, cell.column))

        if not populated:
            sheets.append({"name": worksheet.title, "rows": [], "columns": []})
            continue

        max_row = min(max(row for row, _ in populated), 220)
        max_col = min(max(col for _, col in populated), 60)

        merged_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for merged in worksheet.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = merged.bounds
            if min_row > max_row or min_col > max_col:
                continue
            row_span = min(max_row_m, max_row) - min_row + 1
            col_span = min(max_col_m, max_col) - min_col + 1
            merged_map[(min_row, min_col)] = (row_span, col_span, max_row_m, max_col_m)
            for row in range(min_row, min(max_row_m, max_row) + 1):
                for col in range(min_col, min(max_col_m, max_col) + 1):
                    if (row, col) != (min_row, min_col):
                        covered.add((row, col))

        rows: list[list[dict[str, Any]]] = []
        for row_index in range(1, max_row + 1):
            output_row: list[dict[str, Any]] = []
            for col_index in range(1, max_col + 1):
                if (row_index, col_index) in covered:
                    continue

                cell = worksheet.cell(row_index, col_index)
                row_span, col_span = 1, 1
                if (row_index, col_index) in merged_map:
                    row_span, col_span, _, _ = merged_map[(row_index, col_index)]

                output_row.append(
                    {
                        "row": row_index,
                        "col": col_index,
                        "address": f"{get_column_letter(col_index)}{row_index}",
                        "value": value_to_text(cell.value),
                        "rowSpan": row_span,
                        "colSpan": col_span,
                        "style": cell_style(cell),
                    }
                )
            rows.append(output_row)

        columns = [get_column_letter(col_index) for col_index in range(1, max_col + 1)]
        sheets.append({"name": worksheet.title, "rows": rows, "columns": columns})

    return {"sheets": sheets}


def xls_value_to_text(book: Any, cell: Any) -> str:
    if cell.ctype == 0:
        return ""
    if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, book.datemode).strftime("%d.%m.%Y")
        except Exception:
            return str(cell.value).strip()
    if isinstance(cell.value, float) and cell.value.is_integer():
        return str(int(cell.value))
    return str(cell.value).strip()


def parse_legacy_workbook(path: Path) -> dict[str, Any]:
    if xlrd is None:
        raise ApiError("Не установлен пакет xlrd. Установите: python -m pip install xlrd", 500)

    try:
        book = xlrd.open_workbook(path, formatting_info=True)
    except NotImplementedError:
        book = xlrd.open_workbook(path)

    sheets: list[dict[str, Any]] = []
    for sheet in book.sheets():
        populated: list[tuple[int, int]] = []
        for row_index in range(sheet.nrows):
            for col_index in range(sheet.ncols):
                if xls_value_to_text(book, sheet.cell(row_index, col_index)):
                    populated.append((row_index + 1, col_index + 1))

        if not populated:
            sheets.append({"name": sheet.name, "rows": [], "columns": []})
            continue

        max_row = min(max(row for row, _ in populated), 220)
        max_col = min(max(col for _, col in populated), 60)

        merged_map: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for row_low, row_high, col_low, col_high in getattr(sheet, "merged_cells", []):
            min_row = row_low + 1
            max_row_m = row_high
            min_col = col_low + 1
            max_col_m = col_high
            if min_row > max_row or min_col > max_col:
                continue
            row_span = min(max_row_m, max_row) - min_row + 1
            col_span = min(max_col_m, max_col) - min_col + 1
            merged_map[(min_row, min_col)] = (row_span, col_span)
            for row in range(min_row, min(max_row_m, max_row) + 1):
                for col in range(min_col, min(max_col_m, max_col) + 1):
                    if (row, col) != (min_row, min_col):
                        covered.add((row, col))

        rows: list[list[dict[str, Any]]] = []
        for row_index in range(1, max_row + 1):
            output_row: list[dict[str, Any]] = []
            for col_index in range(1, max_col + 1):
                if (row_index, col_index) in covered:
                    continue
                row_span, col_span = merged_map.get((row_index, col_index), (1, 1))
                cell = sheet.cell(row_index - 1, col_index - 1)
                output_row.append(
                    {
                        "row": row_index,
                        "col": col_index,
                        "address": f"{get_column_letter(col_index) if get_column_letter else col_index}{row_index}",
                        "value": xls_value_to_text(book, cell),
                        "rowSpan": row_span,
                        "colSpan": col_span,
                        "style": {},
                    }
                )
            rows.append(output_row)

        columns = [
            get_column_letter(col_index) if get_column_letter else str(col_index)
            for col_index in range(1, max_col + 1)
        ]
        sheets.append({"name": sheet.name, "rows": rows, "columns": columns})

    return {"sheets": sheets}


TEACHER_RE = re.compile(
    r"\b[А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?\s+[А-ЯЁ]\.?\s*[А-ЯЁ]\.?",
    re.U,
)
TIME_RE = re.compile(r"^\d{1,2}:\d{2}")


def flatten_tree_files(node: dict[str, Any], trail: list[str] | None = None) -> list[dict[str, Any]]:
    trail = trail or []
    if node.get("type") == "file":
        return [{**node, "trail": trail}]

    files: list[dict[str, Any]] = []
    for child in node.get("children", []):
        files.extend(flatten_tree_files(child, [*trail, node.get("title", "")]))
    return files


def file_path_parts(file: dict[str, Any], root_title: str) -> list[str]:
    return [part for part in file.get("trail", []) if part and part != root_title]


def matches_filter(file: dict[str, Any], root_title: str, query: dict[str, list[str]]) -> bool:
    parts = file_path_parts(file, root_title)
    faculty = query.get("faculty", [""])[0]
    semester = query.get("semester", [""])[0]
    section = query.get("section", [""])[0]

    if faculty and (len(parts) < 1 or parts[0] != faculty):
        return False
    if semester and (len(parts) < 2 or parts[1] != semester):
        return False
    if section and (len(parts) < 3 or parts[2] != section):
        return False
    return file.get("extension", "").lower() in {"xlsx", "xlsm", "xltx", "xltm", "xls"}


def clean_value(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_links(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"https?://\S+", "", text)
    text = re.sub(r"\S+@\S+", "", text)
    text = re.sub(r"\b\S+\.(?:ru|com|рф|net|org)\S*", "", text, flags=re.I)
    return clean_value(text)


def normalize_room(value: str) -> str:
    room = strip_links(value)
    if re.search(r"\b(?:идентификатор|код\s+конференции|конференц|парол[ья]?|подключени[ея])\b", room, re.I):
        return ""
    room = re.sub(r"\*.*$", "", room)
    room = re.sub(r"\bссылк\w*.*$", "", room, flags=re.I)
    room = re.sub(r"\([^)]*\)", "", room)
    room = re.sub(r"\bс\s*\d{1,2}\s*[-–]\s*\d{1,2}\b", "", room, flags=re.I)
    room = re.sub(r"^[\s.,;:/\\|-]+", "", room)
    room = re.sub(r"\b(?:аудитория|ауд)\.?\s*", "", room, flags=re.I)
    room = re.sub(r"\b(\d{2,4})\s+([а-яё])\b", r"\1\2", room, flags=re.I)
    room = re.sub(r"(\d)(м\.?\s*к\.?|мод|спо|кл|южн|зоот|общ)\b", r"\1 \2", room, flags=re.I)
    room = re.sub(r"м\.?\s*к\.?", "м.к.", room, flags=re.I)
    room = re.sub(r"\bспо\b", "СПО", room, flags=re.I)
    room = re.sub(r"\bкл\b", "кл", room, flags=re.I)
    room = re.sub(r"\bмод\b", "мод", room, flags=re.I)
    room = re.sub(r"\bюжн\b", "южн", room, flags=re.I)
    room = re.sub(r"\bзоот\.?", "зоот", room, flags=re.I)
    room = re.sub(r"\bобщ\.?", "общ", room, flags=re.I)
    room = re.sub(r"\.{2,}", ".", room)
    return clean_value(room.strip(" ,;:/\\|-"))


def useful_sheet_rows(rows: list[list[dict[str, Any]]]) -> list[list[dict[str, Any]]]:
    useful_rows: list[list[dict[str, Any]]] = []
    used_columns: set[int] = set()

    for row in rows:
        if not any(clean_value(cell.get("value")) for cell in row):
            continue
        useful_rows.append(row)
        for cell in row:
            if clean_value(cell.get("value")):
                start = int(cell.get("col", 0))
                span = int(cell.get("colSpan", 1))
                used_columns.update(range(start, start + span))

    if not used_columns:
        return []

    max_col = max(used_columns)
    return [[cell for cell in row if int(cell.get("col", 0)) <= max_col] for row in useful_rows]


def extract_title(rows: list[list[dict[str, Any]]]) -> str:
    for cell in rows[0] if rows else []:
        value = clean_value(cell.get("value"))
        if value:
            return value
    return ""


def covers_column(cell: dict[str, Any], col: int) -> bool:
    start = int(cell.get("col", 0))
    span = int(cell.get("colSpan", 1))
    return start <= col < start + span


def is_lesson_like_label(value: str) -> bool:
    text = clean_value(value)
    return bool(
        not text
        or len(text) > 48
        or "http" in text.lower()
        or TEACHER_RE.search(text)
        or re.search(r"-(?:Лекц|Лаб|Сем|Прак|Пр\.?|Зач|Экз)", text, re.I)
    )


def compact_group_labels(labels: list[str]) -> str:
    clean_labels: list[str] = []
    for label in labels:
        label = strip_links(label)
        if is_lesson_like_label(label):
            continue
        if label and label not in clean_labels:
            clean_labels.append(label)

    if not clean_labels:
        return "Группа не указана"
    if len(clean_labels) > 6 or len(", ".join(clean_labels)) > 140:
        visible = ", ".join(clean_labels[:4])
        tail = len(clean_labels) - 4
        return f"Несколько групп: {visible}" + (f" и еще {tail}" if tail > 0 else "")
    return ", ".join(clean_labels)


def groups_for_cell(cell: dict[str, Any], groups_by_col: dict[int, str]) -> str:
    labels: list[str] = []
    start = int(cell.get("col", 0))
    span = int(cell.get("colSpan", 1))
    for col in range(start, start + span):
        label = groups_by_col.get(col)
        if label and label not in labels:
            labels.append(label)
    if span > 12 and len(labels) > 8:
        return "Несколько групп"
    return compact_group_labels(labels)


def is_room_text(value: str) -> bool:
    text = normalize_room(value)
    if not text:
        return False
    if TEACHER_RE.search(text):
        return False
    if re.search(r"-(?:Лекц|Лаб|Сем|Прак|Пр\.?|Зач|Экз)", text, re.I):
        return False
    return bool(
        re.search(r"\b\d{2,4}\s*[а-яё.\s-]*(?:м\.?\s*к\.?|мод|корп|ауд)?\b", text, re.I)
        or re.search(r"\b(?:спортзал|актовый|онлайн|дистанц|кабинет)\b", text, re.I)
    )


def split_lesson(raw: str) -> dict[str, Any]:
    raw = str(raw or "")
    lines = [strip_links(line) for line in raw.splitlines() if strip_links(line)]
    teachers: list[str] = []
    for match in TEACHER_RE.finditer(raw):
        teacher = clean_value(match.group(0).replace(" .", "."))
        if teacher not in teachers:
            teachers.append(teacher)

    subject_lines: list[str] = []
    room_lines: list[str] = []

    for index, line in enumerate(lines):
        line_without_teachers = TEACHER_RE.sub("", line)
        line_without_links = strip_links(line_without_teachers).strip(" ,;")
        if not line_without_links:
            continue

        if index == 0:
            subject_lines.append(line_without_links)
        elif is_room_text(line_without_links):
            room_lines.append(line_without_links)
        elif not TEACHER_RE.search(line):
            if not subject_lines:
                subject_lines.append(line_without_links)

    subject = clean_value(" ".join(subject_lines)) or clean_value(lines[0] if lines else raw)
    if len(subject) > 120:
        match = re.search(r"([А-ЯЁA-Zа-яёa-z0-9 .]+-(?:Лекц|Лаб|Сем|Прак|Пр\.?|Зач|Экз))", subject, re.I)
        if match:
            subject = clean_value(match.group(1))
    if len(subject) > 160:
        subject = subject[:157].rstrip() + "..."
    rooms = []
    for room in room_lines:
        room = normalize_room(room)
        if room and room not in rooms:
            rooms.append(room)

    return {
        "subject": subject,
        "teachers": teachers,
        "rooms": rooms,
        "raw": raw,
    }


def schedule_lessons_from_workbook(workbook: dict[str, Any], file: dict[str, Any], root_title: str) -> list[dict[str, Any]]:
    lessons: list[dict[str, Any]] = []
    path_parts = file_path_parts(file, root_title)

    for sheet in workbook.get("sheets", []):
        rows = useful_sheet_rows(sheet.get("rows", []))
        if not rows:
            continue

        title = extract_title(rows)
        body_rows = rows[1:] if title else rows
        data_start = next(
            (
                index
                for index, row in enumerate(body_rows)
                if any(int(cell.get("col", 0)) == 2 and TIME_RE.match(clean_value(cell.get("value"))) for cell in row)
            ),
            -1,
        )
        if data_start < 1:
            continue

        header_rows = body_rows[:data_start]
        data_rows = body_rows[data_start:]
        max_col = max(
            int(cell.get("col", 0)) + int(cell.get("colSpan", 1)) - 1
            for row in body_rows
            for cell in row
        )

        groups_by_col: dict[int, str] = {}
        for col in range(3, max_col + 1):
            labels = [
                strip_links(cell.get("value"))
                for row in header_rows
                for cell in row
                if covers_column(cell, col)
                and strip_links(cell.get("value"))
                and not is_lesson_like_label(strip_links(cell.get("value")))
            ]
            groups_by_col[col] = " / ".join(dict.fromkeys(labels)) or ""

        current_day = ""
        current_time = ""
        current_time_row = 0

        for row in data_rows:
            day = next((clean_value(cell.get("value")) for cell in row if int(cell.get("col", 0)) == 1 and clean_value(cell.get("value"))), "")
            if day:
                current_day = day

            time_value = next((clean_value(cell.get("value")) for cell in row if int(cell.get("col", 0)) == 2 and clean_value(cell.get("value"))), "")
            if time_value:
                current_time = time_value
                current_time_row = 1
            elif current_time:
                current_time_row += 1

            week = 1 if current_time_row <= 1 else 2
            for cell in row:
                if int(cell.get("col", 0)) < 3 or not clean_value(cell.get("value")):
                    continue

                parsed = split_lesson(str(cell.get("value") or ""))
                lesson = {
                    "day": current_day or "День не указан",
                    "time": current_time,
                    "week": week,
                    "weekLabel": "числитель" if week == 1 else "знаменатель",
                    "groups": groups_for_cell(cell, groups_by_col),
                    "subject": parsed["subject"],
                    "teachers": parsed["teachers"],
                    "rooms": parsed["rooms"],
                    "raw": parsed["raw"],
                    "fileTitle": file.get("title", ""),
                    "fileId": file.get("id", ""),
                    "fileUrl": file.get("viewUrl") or file.get("url"),
                    "faculty": path_parts[0] if len(path_parts) > 0 else "",
                    "semester": path_parts[1] if len(path_parts) > 1 else "",
                    "section": path_parts[2] if len(path_parts) > 2 else "",
                    "sheet": sheet.get("name", ""),
                }
                lessons.append(lesson)

    return lessons


def parse_file_lessons(file: dict[str, Any], root_title: str, force: bool = False) -> dict[str, Any]:
    extension = file.get("extension", "").lower() or "xlsx"
    cache_key = hashlib.sha256(
        f"{PARSER_VERSION}:{file.get('id')}:{file.get('modified')}:{extension}".encode("utf-8")
    ).hexdigest()
    lesson_cache_file = LESSON_CACHE_DIR / f"{cache_key}.json"
    if not force and lesson_cache_file.exists():
        return json.loads(lesson_cache_file.read_text(encoding="utf-8"))

    cached = get_cached_file(file["id"], extension, force=force)
    file_head = cached["path"].read_bytes()[:4]
    workbook = (
        parse_legacy_workbook(cached["path"])
        if extension == "xls" and not file_head.startswith(b"PK")
        else parse_workbook(cached["path"])
    )
    result = {
        "file": file,
        "lessons": schedule_lessons_from_workbook(workbook, file, root_title),
    }
    lesson_cache_file.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
    return result


def aggregate_lessons(query: dict[str, list[str]]) -> dict[str, Any]:
    tree = refresh_tree(force=query.get("refreshTree", ["0"])[0] == "1")
    root_title = tree.get("title", "")
    force_files = query.get("refreshFiles", ["0"])[0] == "1"
    files = [file for file in flatten_tree_files(tree) if matches_filter(file, root_title, query)]

    if not files:
        return {"lessons": [], "teachers": [], "rooms": [], "files": [], "errors": []}
    if len(files) > 250:
        raise ApiError("Слишком много файлов для разбора за один раз. Выберите семестр или раздел.", 422)

    cache_key_payload = {
        "parser": PARSER_VERSION,
        "faculty": query.get("faculty", [""])[0],
        "semester": query.get("semester", [""])[0],
        "section": query.get("section", [""])[0],
        "files": sorted(f"{file.get('id')}:{file.get('modified')}" for file in files),
    }
    cache_key = hashlib.sha256(json.dumps(cache_key_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    aggregate_cache_file = AGGREGATE_CACHE_DIR / f"{cache_key}.json"
    if (
        not force_files
        and query.get("refreshTree", ["0"])[0] != "1"
        and aggregate_cache_file.exists()
        and time.time() - aggregate_cache_file.stat().st_mtime < AGGREGATE_CACHE_SECONDS
    ):
        return json.loads(aggregate_cache_file.read_text(encoding="utf-8"))

    lessons: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    workers = min(4, len(files))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(parse_file_lessons, file, root_title, force_files): file for file in files}
        for future in as_completed(futures):
            file = futures[future]
            try:
                result = future.result()
                lessons.extend(result["lessons"])
            except Exception as exc:
                errors.append({"file": file.get("title", ""), "error": str(exc)})

    teacher_names = sorted(
        {teacher for lesson in lessons for teacher in lesson["teachers"]},
        key=lambda value: value.lower(),
    )
    room_names = sorted(
        {room for lesson in lessons for room in lesson["rooms"]},
        key=lambda value: value.lower(),
    )

    result = {
        "lessons": lessons,
        "teachers": teacher_names,
        "rooms": room_names,
        "files": [{"id": file.get("id"), "title": file.get("title"), "path": file_path_parts(file, root_title)} for file in files],
        "errors": errors,
        "refreshedAt": now_iso(),
    }
    aggregate_cache_file.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
    return result


def api_tree(force: bool) -> dict[str, Any]:
    data = refresh_tree(force=force)
    return {
        "rootFolderId": ROOT_FOLDER_ID,
        "tree": data,
        "cached": not force,
        "error": tree_cache.get("error"),
    }


def api_schedule(query: dict[str, list[str]], file_id: str) -> dict[str, Any]:
    extension = query.get("ext", ["xlsx"])[0].lower() or "xlsx"
    title = query.get("title", ["Расписание"])[0]
    force = query.get("refresh", ["0"])[0] == "1"

    if extension not in {"xlsx", "xlsm", "xltx", "xltm", "xls"}:
        return {
            "id": file_id,
            "title": title,
            "extension": extension,
            "viewUrl": f"https://drive.google.com/file/d/{file_id}/view",
            "downloadUrl": drive_download_url(file_id),
            "message": "Этот тип файла лучше открыть в Google Drive.",
        }

    cached = get_cached_file(file_id, extension, force=force)
    file_head = cached["path"].read_bytes()[:4]
    workbook = (
        parse_legacy_workbook(cached["path"])
        if extension == "xls" and not file_head.startswith(b"PK")
        else parse_workbook(cached["path"])
    )
    return {
        "id": file_id,
        "title": title,
        "extension": extension,
        "viewUrl": f"https://drive.google.com/file/d/{file_id}/view",
        "downloadUrl": drive_download_url(file_id),
        "size": cached["size"],
        "sha256": cached["sha256"],
        "cachedAt": cached["cachedAt"],
        **workbook,
    }


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def log_message(self, format: str, *args: Any) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), format % args))

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)

        try:
            if parsed.path == "/api/tree":
                self.send_json(api_tree(force=query.get("refresh", ["0"])[0] == "1"))
                return

            if parsed.path == "/api/aggregate":
                self.send_json(aggregate_lessons(query))
                return

            match = re.fullmatch(r"/api/schedule/([^/]+)", parsed.path)
            if match:
                self.send_json(api_schedule(query, match.group(1)))
                return

            if parsed.path == "/health":
                self.send_json({"ok": True, "time": now_iso()})
                return

            if parsed.path == "/":
                self.path = "/index.html"
            return super().do_GET()
        except ApiError as exc:
            self.send_json({"error": str(exc)}, status=exc.status)
        except urllib.error.URLError as exc:
            self.send_json({"error": f"Ошибка сети при обращении к Google Drive: {exc}"}, status=502)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)


def main() -> None:
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8000
    threading.Thread(target=background_refresh, daemon=True).start()
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"VSAU schedule site: http://127.0.0.1:{port}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
