from __future__ import annotations

import hashlib
import io
import json
import mimetypes
import re
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    load_workbook = None
    get_column_letter = None

try:
    import xlrd
except ModuleNotFoundError:
    xlrd = None


ROOT_DIR = Path(__file__).resolve().parent
STATIC_DIR = ROOT_DIR / "static"
CACHE_DIR = ROOT_DIR / "data" / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

ROOT_FOLDER_ID = "1Gn8OEzbtxFBusuCnPCTBjQoK1AWKiVN2"
TREE_CACHE_SECONDS = 10 * 60
FILE_CACHE_SECONDS = 5 * 60
POLL_SECONDS = 10 * 60
MAX_TREE_DEPTH = 6

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
)

tree_lock = threading.Lock()
tree_cache: dict[str, Any] = {"updated_at": 0.0, "data": None, "error": None}


class ApiError(Exception):
    def __init__(self, message: str, status: int = 500) -> None:
        super().__init__(message)
        self.status = status


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def fetch_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=30) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        return response.read().decode(charset, errors="replace")


def download_bytes(url: str) -> tuple[bytes, str]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=60) as response:
        content_type = response.headers.get("content-type", "")
        return response.read(), content_type


def drive_embedded_url(folder_id: str) -> str:
    query = urllib.parse.urlencode({"id": folder_id})
    return f"https://drive.google.com/embeddedfolderview?{query}#list"


def drive_download_url(file_id: str) -> str:
    query = urllib.parse.urlencode({"export": "download", "id": file_id})
    return f"https://drive.google.com/uc?{query}"


def drive_preview_url(file_id: str) -> str:
    return f"https://drive.google.com/file/d/{urllib.parse.quote(file_id)}/preview"


def extract_drive_id(url: str) -> str:
    patterns = (
        r"/drive/folders/([^/?#]+)",
        r"/file/d/([^/?#]+)",
        r"[?&]id=([^&#]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return urllib.parse.unquote(match.group(1))
    raise ApiError(f"Не удалось найти id Google Drive в ссылке: {url}", 422)


def clean_html(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    return urllib.parse.unquote(value).replace("&amp;", "&").strip()


def parse_folder_entries(html: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    pattern = re.compile(
        r'<a href="(?P<href>[^"]+)"[^>]*>.*?'
        r'<div class="flip-entry-title">(?P<title>.*?)</div>.*?'
        r'<div class="flip-entry-last-modified"><div>(?P<modified>.*?)</div>',
        re.S,
    )

    for match in pattern.finditer(html):
        href = match.group("href").replace("&amp;", "&")
        title = clean_html(match.group("title"))
        modified = clean_html(match.group("modified"))
        if not title:
            continue

        is_folder = "/drive/folders/" in href
        item_id = extract_drive_id(href)
        extension = Path(title).suffix.lower().lstrip(".")
        entries.append(
            {
                "id": item_id,
                "title": title,
                "modified": modified,
                "type": "folder" if is_folder else "file",
                "extension": extension,
                "url": href,
                "downloadUrl": drive_download_url(item_id) if not is_folder else None,
                "viewUrl": href,
            }
        )

    return entries


def fetch_folder(folder_id: str) -> dict[str, Any]:
    html = fetch_text(drive_embedded_url(folder_id))
    title_match = re.search(r"<title>(.*?)</title>", html, re.S)
    title = clean_html(title_match.group(1)) if title_match else "Google Drive"
    return {
        "id": folder_id,
        "title": title.replace(" - Google Drive", "").replace(" – Dysk Google", ""),
        "type": "folder",
        "url": f"https://drive.google.com/drive/folders/{folder_id}",
        "children": parse_folder_entries(html),
    }


def build_tree(folder_id: str, depth: int = 0) -> dict[str, Any]:
    folder = fetch_folder(folder_id)
    if depth >= MAX_TREE_DEPTH:
        return folder

    children: list[dict[str, Any]] = list(folder["children"])
    folder_indexes = [index for index, child in enumerate(children) if child["type"] == "folder"]
    if folder_indexes:
        workers = min(8, len(folder_indexes))
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(build_tree, children[index]["id"], depth + 1): index
                for index in folder_indexes
            }
            for future in as_completed(futures):
                index = futures[future]
                child = children[index]
                try:
                    nested = future.result()
                    nested["title"] = child["title"]
                    nested["modified"] = child.get("modified")
                    children[index] = nested
                except Exception as exc:
                    child["error"] = str(exc)
                    children[index] = child

    folder["children"] = children
    return folder


def refresh_tree(force: bool = False) -> dict[str, Any]:
    with tree_lock:
        age = time.time() - tree_cache["updated_at"]
        if not force and tree_cache["data"] is not None and age < TREE_CACHE_SECONDS:
            return tree_cache["data"]

        try:
            data = build_tree(ROOT_FOLDER_ID)
            data["refreshedAt"] = now_iso()
            tree_cache.update({"updated_at": time.time(), "data": data, "error": None})
            return data
        except Exception as exc:
            tree_cache["error"] = str(exc)
            if tree_cache["data"] is not None:
                return tree_cache["data"]
            raise


def background_refresh() -> None:
    while True:
        try:
            refresh_tree(force=True)
        except Exception:
            pass
        time.sleep(POLL_SECONDS)


def cache_file_path(file_id: str, extension: str) -> Path:
    extension = extension or "bin"
    return CACHE_DIR / f"{file_id}.{extension}"


def get_cached_file(file_id: str, extension: str, force: bool = False) -> dict[str, Any]:
    path = cache_file_path(file_id, extension)
    age = time.time() - path.stat().st_mtime if path.exists() else float("inf")
    if force or not path.exists() or age > FILE_CACHE_SECONDS:
        data, content_type = download_bytes(drive_download_url(file_id))
        if b"<html" in data[:300].lower() and "text/html" in content_type:
            raise ApiError("Google Drive вернул HTML-страницу вместо файла. Проверьте публичный доступ.", 502)
        path.write_bytes(data)

    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "path": path,
        "size": path.stat().st_size,
        "sha256": digest,
        "cachedAt": datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat(timespec="seconds"),
    }


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def cell_style(cell: Any) -> dict[str, Any]:
    fill = getattr(cell.fill.fgColor, "rgb", None)
    if fill and len(fill) == 8 and fill != "00000000":
        fill = "#" + fill[2:]
    else:
        fill = ""

    color = getattr(cell.font.color, "rgb", None)
    if color and isinstance(color, str) and len(color) == 8 and color != "00000000":
        color = "#" + color[2:]
    else:
        color = ""

    return {
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "fill": fill,
        "color": color,
        "align": cell.alignment.horizontal or "",
        "valign": cell.alignment.vertical or "",
    }


def parse_workbook(path: Path) -> dict[str, Any]:
    if load_workbook is None:
        raise ApiError("Не установлен пакет openpyxl. Установите: python -m pip install openpyxl", 500)

    source: Any = io.BytesIO(path.read_bytes()) if path.suffix.lower() == ".xls" else path
    workbook = load_workbook(source, data_only=True)
    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        populated: list[tuple[int, int]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if value_to_text(cell.value):
                    populated.append((cell.row, cell.column))

        if not populated:
            sheets.append({"name": worksheet.title, "rows": [], "columns": []})
            continue

        max_row = min(max(row for row, _ in populated), 220)
        max_col = min(max(col for _, col in populated), 60)

        merged_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for merged in worksheet.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = merged.bounds
            if min_row > max_row or min_col > max_col:
                continue
            row_span = min(max_row_m, max_row) - min_row + 1
            col_span = min(max_col_m, max_col) - min_col + 1
            merged_map[(min_row, min_col)] = (row_span, col_span, max_row_m, max_col_m)
            for row in range(min_row, min(max_row_m, max_row) + 1):
                for col in range(min_col, min(max_col_m, max_col) + 1):
                    if (row, col) != (min_row, min_col):
                        covered.add((row, col))

        rows: list[list[dict[str, Any]]] = []
        for row_index in range(1, max_row + 1):
            output_row: list[dict[str, Any]] = []
            for col_index in range(1, max_col + 1):
                if (row_index, col_index) in covered:
                    continue

                cell = worksheet.cell(row_index, col_index)
                row_span, col_span = 1, 1
                if (row_index, col_index) in merged_map:
                    row_span, col_span, _, _ = merged_map[(row_index, col_index)]

                output_row.append(
                    {
                        "row": row_index,
                        "col": col_index,
                        "address": f"{get_column_letter(col_index)}{row_index}",
                        "value": value_to_text(cell.value),
                        "rowSpan": row_span,
                        "colSpan": col_span,
                        "style": cell_style(cell),
                    }
                )
            rows.append(output_row)

        columns = [get_column_letter(col_index) for col_index in range(1, max_col + 1)]
        sheets.append({"name": worksheet.title, "rows": rows, "columns": columns})

    return {"sheets": sheets}


def xls_value_to_text(book: Any, cell: Any) -> str:
    if cell.ctype == 0:
        return ""
    if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, book.datemode).strftime("%d.%m.%Y")
        except Exception:
            return str(cell.value).strip()
    if isinstance(cell.value, float) and cell.value.is_integer():
        return str(int(cell.value))
    return str(cell.value).strip()


def parse_legacy_workbook(path: Path) -> dict[str, Any]:
    if xlrd is None:
        raise ApiError("Не установлен пакет xlrd. Установите: python -m pip install xlrd", 500)

    try:
        book = xlrd.open_workbook(path, formatting_info=True)
    except NotImplementedError:
        book = xlrd.open_workbook(path)

    sheets: list[dict[str, Any]] = []
    for sheet in book.sheets():
        populated: list[tuple[int, int]] = []
        for row_index in range(sheet.nrows):
            for col_index in range(sheet.ncols):
                if xls_value_to_text(book, sheet.cell(row_index, col_index)):
                    populated.append((row_index + 1, col_index + 1))

        if not populated:
            sheets.append({"name": sheet.name, "rows": [], "columns": []})
            continue

        max_row = min(max(row for row, _ in populated), 220)
        max_col = min(max(col for _, col in populated), 60)

        merged_map: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for row_low, row_high, col_low, col_high in getattr(sheet, "merged_cells", []):
            min_row = row_low + 1
            max_row_m = row_high
            min_col = col_low + 1
            max_col_m = col_high
            if min_row > max_row or min_col > max_col:
                continue
            row_span = min(max_row_m, max_row) - min_row + 1
            col_span = min(max_col_m, max_col) - min_col + 1
            merged_map[(min_row, min_col)] = (row_span, col_span)
            for row in range(min_row, min(max_row_m, max_row) + 1):
                for col in range(min_col, min(max_col_m, max_col) + 1):
                    if (row, col) != (min_row, min_col):
                        covered.add((row, col))

        rows: list[list[dict[str, Any]]] = []
        for row_index in range(1, max_row + 1):
            output_row: list[dict[str, Any]] = []
            for col_index in range(1, max_col + 1):
                if (row_index, col_index) in covered:
                    continue
                row_span, col_span = merged_map.get((row_index, col_index), (1, 1))
                cell = sheet.cell(row_index - 1, col_index - 1)
                output_row.append(
                    {
                        "row": row_index,
                        "col": col_index,
                        "address": f"{get_column_letter(col_index) if get_column_letter else col_index}{row_index}",
                        "value": xls_value_to_text(book, cell),
                        "rowSpan": row_span,
                        "colSpan": col_span,
                        "style": {},
                    }
                )
            rows.append(output_row)

        columns = [
            get_column_letter(col_index) if get_column_letter else str(col_index)
            for col_index in range(1, max_col + 1)
        ]
        sheets.append({"name": sheet.name, "rows": rows, "columns": columns})

    return {"sheets": sheets}


def api_tree(force: bool) -> dict[str, Any]:
    data = refresh_tree(force=force)
    return {
        "rootFolderId": ROOT_FOLDER_ID,
        "tree": data,
        "cached": not force,
        "error": tree_cache.get("error"),
    }


def api_schedule(query: dict[str, list[str]], file_id: str) -> dict[str, Any]:
    extension = query.get("ext", ["xlsx"])[0].lower() or "xlsx"
    title = query.get("title", ["Расписание"])[0]
    force = query.get("refresh", ["0"])[0] == "1"

    if extension not in {"xlsx", "xlsm", "xltx", "xltm", "xls"}:
        inline_extensions = {"pdf", "png", "jpg", "jpeg", "gif", "webp", "svg", "txt", "csv", "html", "htm"}
        preview_kind = (
            "pdf" if extension == "pdf"
            else "image" if extension in {"png", "jpg", "jpeg", "gif", "webp", "svg"}
            else "text" if extension in {"txt", "csv", "html", "htm"}
            else "google"
        )
        preview_params = urllib.parse.urlencode({"ext": extension, "title": title})
        return {
            "id": file_id,
            "title": title,
            "extension": extension,
            "viewUrl": f"https://drive.google.com/file/d/{file_id}/view",
            "downloadUrl": drive_download_url(file_id),
            "previewKind": preview_kind,
            "previewUrl": f"/api/file/{file_id}?{preview_params}" if extension in inline_extensions else drive_preview_url(file_id),
            "message": "Файл можно просмотреть прямо на сайте.",
        }

    cached = get_cached_file(file_id, extension, force=force)
    file_head = cached["path"].read_bytes()[:4]
    workbook = (
        parse_legacy_workbook(cached["path"])
        if extension == "xls" and not file_head.startswith(b"PK")
        else parse_workbook(cached["path"])
    )
    return {
        "id": file_id,
        "title": title,
        "extension": extension,
        "viewUrl": f"https://drive.google.com/file/d/{file_id}/view",
        "downloadUrl": drive_download_url(file_id),
        "size": cached["size"],
        "sha256": cached["sha256"],
        "cachedAt": cached["cachedAt"],
        **workbook,
    }


def send_cached_file(handler: SimpleHTTPRequestHandler, query: dict[str, list[str]], file_id: str) -> None:
    extension = query.get("ext", ["bin"])[0].lower() or "bin"
    title = query.get("title", [f"{file_id}.{extension}"])[0]
    force = query.get("refresh", ["0"])[0] == "1"
    cached = get_cached_file(file_id, extension, force=force)
    content_type = mimetypes.guess_type(title)[0] or mimetypes.guess_type(f"file.{extension}")[0] or "application/octet-stream"
    data = cached["path"].read_bytes()

    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Disposition", f"inline; filename*=UTF-8''{urllib.parse.quote(title)}")
    handler.send_header("Cache-Control", "private, max-age=300")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def log_message(self, format: str, *args: Any) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), format % args))

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        query = urllib.parse.parse_qs(parsed.query)

        try:
            if parsed.path == "/api/tree":
                self.send_json(api_tree(force=query.get("refresh", ["0"])[0] == "1"))
                return

            match = re.fullmatch(r"/api/schedule/([^/]+)", parsed.path)
            if match:
                self.send_json(api_schedule(query, match.group(1)))
                return

            match = re.fullmatch(r"/api/file/([^/]+)", parsed.path)
            if match:
                send_cached_file(self, query, match.group(1))
                return

            if parsed.path == "/health":
                self.send_json({"ok": True, "time": now_iso()})
                return

            if parsed.path == "/":
                self.path = "/index.html"
            return super().do_GET()
        except ApiError as exc:
            self.send_json({"error": str(exc)}, status=exc.status)
        except urllib.error.URLError as exc:
            self.send_json({"error": f"Ошибка сети при обращении к Google Drive: {exc}"}, status=502)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)


def main() -> None:
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8000
    threading.Thread(target=background_refresh, daemon=True).start()
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"VSAU schedule site: http://127.0.0.1:{port}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
